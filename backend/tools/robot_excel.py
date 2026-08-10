"""Excel import/export helpers for robot funnel cards."""
from io import BytesIO

from openpyxl import Workbook, load_workbook

ROBOT_HEADERS = [
    'nickname',
    'age',
    'job',
    'city',
    'bio',
    'photo_urls',
    'tags',
    'mbti',
    'zodiac',
    'relationship',
    'country',
    'locale',
    'sort_order',
    'is_traveling',
]

# Accept Chinese aliases in the header row
HEADER_ALIASES = {
    'nickname': {'nickname', '昵称', 'name', '名字'},
    'age': {'age', '年龄'},
    'job': {'job', '职业', '工作'},
    'city': {'city', '城市'},
    'bio': {'bio', '简介', '签名', 'about'},
    'photo_urls': {'photo_urls', 'photos', 'photo', '头像', '照片', '图片链接'},
    'tags': {'tags', '标签', '兴趣'},
    'mbti': {'mbti'},
    'zodiac': {'zodiac', '星座'},
    'relationship': {'relationship', '感情状态', '关系'},
    'country': {'country', 'region', '地区', '国家'},
    'locale': {'locale', 'language', '语言'},
    'sort_order': {'sort_order', 'sort', '排序'},
    'is_traveling': {'is_traveling', 'traveling', '旅行中'},
}


def _cell_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_list(value):
    text = _cell_str(value)
    if not text:
        return []
    for sep in ('|', ';', '；', '\n'):
        text = text.replace(sep, ',')
    return [p.strip() for p in text.split(',') if p.strip()]


def _parse_bool(value, default=False):
    text = _cell_str(value).lower()
    if not text:
        return default
    if text in ('1', 'true', 'yes', 'y', '是', '真'):
        return True
    if text in ('0', 'false', 'no', 'n', '否', '假'):
        return False
    return default


def _parse_int(value, default=0):
    text = _cell_str(value)
    if not text:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def build_robot_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'robots'
    ws.append(ROBOT_HEADERS)
    ws.append([
        'Nina', 24, 'UX Designer', 'California',
        'I love museum and stand-up comedy.',
        'https://images.unsplash.com/photo-1524504388940-b1c1722653e1?w=800',
        'Travel,Art', 'ENTP', 'Virgo', 'Open Relationship',
        '*', 'en', 0, 0,
    ])
    ws.append([
        'Ava', 26, 'Product Manager', 'Tokyo',
        'Looking for sincere connections.',
        'https://images.unsplash.com/photo-1494790108377-be9c29b29330?w=800|https://images.unsplash.com/photo-1438761681033-6461ffad8d80?w=800',
        'Coffee,Hiking', 'INFJ', 'Leo', '',
        'JP', 'ja', 1, 1,
    ])
    # column widths
    widths = {
        'A': 14, 'B': 8, 'C': 16, 'D': 14, 'E': 36, 'F': 48, 'G': 18,
        'H': 10, 'I': 10, 'J': 18, 'K': 10, 'L': 10, 'M': 10, 'N': 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    hint = wb.create_sheet('readme')
    hint.append(['Field', 'Required', 'Notes'])
    hint.append(['nickname', 'yes', 'Display name'])
    hint.append(['age', 'no', 'Integer, default 24'])
    hint.append(['job / city / bio', 'no', 'Text'])
    hint.append(['photo_urls', 'no', 'Comma or | separated image URLs'])
    hint.append(['tags', 'no', 'Comma separated'])
    hint.append(['country', 'no', '* / US / JP / CN ... default *'])
    hint.append(['locale', 'no', 'en / zh / ja ... default en'])
    hint.append(['sort_order', 'no', 'Integer, default 0'])
    hint.append(['is_traveling', 'no', '0/1 or true/false'])
    hint.append(['', '', 'Chinese header aliases are also accepted.'])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _map_headers(header_row):
    mapping = {}
    for idx, raw in enumerate(header_row):
        key = _cell_str(raw).lower()
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if key in {a.lower() for a in aliases}:
                mapping[field] = idx
                break
    return mapping


def parse_robot_import_rows(file_obj, default_country='*', default_locale='en'):
    """
    Parse uploaded xlsx. Returns (rows, errors).
    Each row is a dict ready for FunnelPool.objects.create(**row) minus app_id/pool.
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ['empty file']

    mapping = _map_headers(header)
    if 'nickname' not in mapping:
        return [], ['missing required column: nickname']

    rows = []
    errors = []
    for line_no, values in enumerate(rows_iter, start=2):
        if not values or all(v is None or str(v).strip() == '' for v in values):
            continue
        def get(field, default=''):
            idx = mapping.get(field)
            if idx is None or idx >= len(values):
                return default
            return values[idx]

        nickname = _cell_str(get('nickname'))
        if not nickname:
            errors.append(f'row {line_no}: nickname required')
            continue

        country = _cell_str(get('country')) or default_country or '*'
        locale = (_cell_str(get('locale')) or default_locale or 'en').lower()
        rows.append({
            'nickname': nickname[:64],
            'age': _parse_int(get('age'), 24),
            'job': _cell_str(get('job'))[:64],
            'city': _cell_str(get('city'))[:64],
            'bio': _cell_str(get('bio')),
            'photo_urls': _split_list(get('photo_urls')),
            'tags': _split_list(get('tags')),
            'mbti': _cell_str(get('mbti'))[:8],
            'zodiac': _cell_str(get('zodiac'))[:16],
            'relationship': _cell_str(get('relationship'))[:64],
            'country': country[:16] or '*',
            'locale': locale[:16] or 'en',
            'sort_order': _parse_int(get('sort_order'), 0),
            'is_traveling': _parse_bool(get('is_traveling'), False),
        })
    return rows, errors
